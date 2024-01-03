import sys
import json
import os
import random
import math
from datetime import date, timedelta, datetime
import pickle
import re

import sys
sys.path.append("..")
from api.settings.imports import *
from api.settings.utils import *

from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font

import csv
import pandas as pd

export_records = []

# profiles
filters = {}
profiles = get_profiles(app_id, filters=filters)
for profile in profiles:
   print(profile)
   export_records.append(profile)


# Exporting to Excel

#Creating Excel spreadsheet
book = Workbook()
records = book.create_sheet("Records")
records.title = "Records"

records.cell(1,1).value = 'profile Id'
records.cell(1,2).value = 'Name'
records.cell(1,3).value = 'Something Else'

current_row = 1
for item, record in enumerate(export_records):

    profile_id = record['profile_id']
    name = record['name']
    something_else = 'Something else ' + randomstr()

    # Export the record
    current_row = current_row + 1
    row = current_row
    records.cell(row=row, column=1).value = profile_id
    records.cell(row=row, column=2).value = name
    records.cell(row=row, column=3).value = something_else


# Save the excel workbook file to memory
export_name = 'Export_' + randomstr() + '.xlsx'
book.save(export_name)

print('Exported Excel')


# Exporting to CSV

# Headers
headers = ['profile Id', 'Name', 'Something Else']

# File name
csv_export_name = 'Export_' + randomstr() + '.csv'

# Write to CSV
with open(csv_export_name, mode='w', newline='', encoding='utf-8') as file:
    writer = csv.DictWriter(file, fieldnames=headers)

    writer.writeheader()
    for record in export_records:
        writer.writerow({
            'profile Id': record['profile_id'],
            'Name': record['name'],
            'Something Else': 'Something else ' + randomstr()
        })

print('Exported CSV')


# Exporting to JSON

# Modify each record for JSON format if necessary
json_records = [{'profile_id': rec['profile_id'], 'name': rec['name'], 'something_else': 'Something else ' + randomstr()} for rec in export_records]

# File name
json_export_name = 'Export_' + randomstr() + '.json'

# Write to JSON
with open(json_export_name, 'w', encoding='utf-8') as file:
    json.dump(json_records, file, ensure_ascii=False, indent=4)

print('Exported JSON')



# Reading JSON data into a pandas DataFrame
df = pd.read_json(json_export_name)

# Printing the DataFrame
print(df)


